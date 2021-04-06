from openpyxl import load_workbook
from django.shortcuts import render
from ..models.model_major import Major
from ..models.model_school import School
from ..models.model_transfer_course import TransferCourse
from ..models.model_major_requirement import MajorRequirement
from ..models.model_transferevaluation import Transferevaluation
from ..models.model_approver import Approver


def import_file(request):
    if request.method == 'POST':
        file = request.FILES['document']
        print(type(file))
        import_data(file)
    return render(request, 'import.html')


def import_data(request):
    wb = load_workbook(filename=request, data_only=True)
    majors = import_major(wb)
    schools = []
    courses = []
    approvers = []
    major_reqs = []
    evals = []

    data = test_get_data_by_major(wb)
    schools.extend(data[0])
    courses.extend(data[1])
    approvers.extend(data[2])
    major_reqs.extend(data[3])
    evals.extend(data[4])
    import_school(schools)
    import_course(courses)
    import_approvers(approvers)
    import_requirement(major_reqs)
    import_evaluations(evals)


def import_major(wb_object):
    '''
    Goes through the majors and adds them into the database
    '''
    major_names = wb_object.sheetnames
    count = 1
    for major in enumerate(major_names):
        major_data = Major(count, major[1])
        major_data.save()
        count = count + 1
    return major_names


def import_school(schools):
    '''
    Goes through the Schools and adds them into the database
    '''
    #print(schools)
    unique_schools = []
    count = 1
    for school in schools:
        if school not in unique_schools:
            unique_schools.append(school)
    for school_name in unique_schools:
        school_data = School(count, school_name, "N/A")
        school_data.save()
        count = count + 1
    #return schools


def import_course(courses):
    '''
    Goes through the courses and adds them into the database
    '''
    count = 1
    course_list = []
    #print(courses, '\n')
    #unique_courses = []
    #for course_lst in courses:
    #    if course_lst[0:2] not in unique_courses:
    #        unique_courses.append(course_lst)
    #print(courses,'This is in course_import')
    for course in courses:
        #print(course)
        if course not in course_list:
            course_data = TransferCourse(count, course[0], course[1], course[2])
            course_data.save()
            count = count + 1


def import_approvers(approvers):
    '''
    Goes through the approvers and adds them into the database
    '''
    count = 1
    unique_approvers = []
    for approver in approvers:
        if approver not in unique_approvers:
            unique_approvers.append(approver)
    for approver in unique_approvers:
        approver_data = Approver(count, approver)
        approver_data.save()
        count = count + 1


def import_requirement(major_reqs):
    '''
    Goes through the requirements and adds them into the database
    '''
    count = 1
    print(major_reqs,'These kskdikdidi')
    for req in major_reqs:
        req_data = MajorRequirement(count, req[1], req[0])
        req_data.save()
        count = count + 1


def import_evaluations(evals):
    '''
    Goes through the evaluations and adds them into the database
    '''
    count = 1
    #print(evals)
    for evaluation in evals:
        #print(evaluation)
        eval_data = Transferevaluation(
            count,
            evaluation[0],  # course_id
            evaluation[1],  # major-req_id
            evaluation[2],  #eval_row[2] sem_year_taken
            evaluation[5],
            evaluation[3],  # approved_status
            evaluation[6],  # comment
            evaluation[4]   # approver_id
        )
        eval_data.save()
        #print(eval_data)
        count = count + 1


def test_get_data_by_major(transfer_wb):
    """
    Tests get_data_by_major()
    """
    schools = []
    courses = []
    approvers = []
    major_reqs = []
    evals = []

    #major_ws = transfer_wb[major_name]
    #major_id = transfer_wb.sheetnames.index(major_name) + 1
    get_data_by_major(
        transfer_wb, schools, courses, approvers, major_reqs, evals
    )
    print(major_reqs,'This is in get_data_by_major')
    return [schools, courses, approvers, major_reqs, evals]


def get_unique_vals_from_col(transfer_wb, col_idx):
    """
    Computes and returns a list of unique values from the cell values of the
    column at col_idx in the worksheet major_ws.
    Used to get School and Approver name attribute values and
    MajorRequirement description attribute values.
    major_ws: worksheet
    col_idx: integer, column index
    Returns: tuple of unique values
    """
    # Accumulator is a set because it let us add ONLY unique values
    value_set = set()
    # Iterate over all the rows in the selected column, col_idx
    # Loop variable is a tuple with one element, call value
    for sheet in transfer_wb:
        #print(major_ws)
        for row_tuple in sheet.iter_rows(
                min_row=2, max_row=sheet.max_row,
                min_col=col_idx, max_col=col_idx,
                values_only=True):
            # Extract element value from the tuple and add it to value_set
            # If element value already exists in value_set, nothing happens
            value_set.add(row_tuple[0])
    #print(value_set,"This is in vjsvh")
    return list(value_set)

def get_unique_reqs_vals_from_col(transfer_wb, col_idx):
    """
    Computes and returns a list of unique values from the cell values of the
    column at col_idx in the worksheet major_ws.
    Used to get School and Approver name attribute values and
    MajorRequirement description attribute values.
    major_ws: worksheet
    col_idx: integer, column index
    Returns: tuple of unique values
    """
    # Accumulator is a set because it let us add ONLY unique values
    #value_set = set()
    major_descriptions = []
    # Iterate over all the rows in the selected column, col_idx
    # Loop variable is a tuple with one element, call value
    for major_ws in transfer_wb:
        major_id = transfer_wb.sheetnames.index(major_ws.title) + 1
        value_set = set()
        for row_tuple in major_ws.iter_rows(
            min_row=2, max_row=major_ws.max_row,
            min_col=col_idx, max_col=col_idx,
            values_only=True):
    # Extract element value from the tuple and add it to value_set
    # If element value already exists in value_set, nothing happe
            value_set.add(row_tuple[0])
        print(value_set, 'this is value_set')
        for major_description in list(value_set):
            major_descriptions.append([major_id, major_description])
    print(major_descriptions, 'major_descriptions')
    return major_descriptions


def eval_with_fk(major_id, eval_row, schools, courses, approvers, major_reqs):
    """
    Uses data from eval_row (12-element list) and IDs from courses, approvers,
    and major_reqs to create and return a list that has foreign key values for
    approver ID, course ID, and major requirement ID.
    major_id: index of worksheet in the workbook incremented by 1
    eval_row: list of values from columns 1 to 12 in a row in a worksheet
    schools, courses, approvers, major_reqs: lists of data
    Returns: list of values representing an eval netry in the
    TransferEvaluation entity with forign keys
        course ID, majore requirement ID, and approver ID
    """
    eval_data = []
    #print(courses)

    # get school name from column 1
    # find school ID in schools list
    school_name = eval_row[0]
    school_id = schools.index(school_name) + 1

    # get subject number and course title from columns 2 and 3
    # find course ID in courses list
    course_data = []
    course_data.append(school_id)
    course_data.extend(eval_row[1:3])
    #print(course_data)
    # print(course_data)
    # course_id = TransferCourse.objects.get(transfer_course_id=transfer_course_id)
    #print(courses,'These are courses')
    #print(course_data,'This is course_data')
    course_id = courses.index(course_data) + 1
    #print(course_id,'This is transfer_eval')
    #print(course_id)
    eval_data.append(course_id)
    #print(eval_data,'after courses')
    # get major requirement description from column 6
    # find major requirement ID in major_reqs list
    major_req_desc = eval_row[5]
    major_req_id = major_reqs.index([major_id, major_req_desc]) + 1
    eval_data.append(major_req_id)
    #print(major_id)
    #print(major_req_id)
    #print(major_id, major_req_id)

    # Get sem & year taken and approved status from columns 4 and 5
    eval_data.extend(eval_row[3:5])

    # get approver name from column 8 and find approver ID in approvers list
    approver_name = eval_row[7]
    approver_id = approvers.index(approver_name) + 1
    eval_data.append(approver_id)

    # get expiration date from column 9
    eval_data.append(eval_row[8])

    # get comment from column 12
    eval_data.append(eval_row[11])
    # print(eval_row, '\n')
    #print(eval_data, '\n')
    return eval_data


def course_with_fk(course_row, schools):
    """
    Uses data from course_row and the ID of of the school name  to create and
    return a list that has the foreign key value for school_id
    course_row: list of values from columns 1 to 3 in a row in a worksheet
    schools: list of strings, representing school names
    Returns:
        list of values representing a course entry in the Course entity with
        foreign key school ID
    """
    #print(schools)
    course_data = []
    #print(course_row,'This in course by Fk')
    school_name = course_row[0]
    school_id = schools.index(school_name) + 1
    course_data.append(school_id)
    course_data.extend(course_row[1:3])
    return course_data
    #print(course_data,'hh')

def get_course_by_major(transfer_wb, start, end, schools):
    """
    Computes and returns a list of 3-element sublists with course data from
    major_ws worksheet. The 3-element sublist has data from columns
        1 (school name), 2 (subject num), and 3 (course title) in a worksheet.
    Replaces school name with school ID in the schools list.
    major_ws: worksheet
    start, end: column indices
    schools: lists of unique names
    Returns: list of 3-elemnt sublists
    """
    courses = []
    course_list = []
    for sheet in transfer_wb:
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=start, max_col=end, values_only=True):
        # print(row,'row in course')
            course_data = course_with_fk(row, schools)
        #print(course_data)
            courses.append(course_data)
    #print(course_list,'This is in courses')
    #for course in course_list:
        #print(course,'This in cousejdfkddk')
    #    if course not in courses:
    #        courses.append(course)
    #print(courses,'jkfuuk')
    return courses


def get_eval_by_major(
        transfer_wb, start, end, schools, courses, approvers,
        major_reqs):
    """
    Computes and returns a list of 7-elemnent sublists with transfer evaluation
    data from major_ws worksheet. The 7-element sublist has data from
        4 columns:
            4 (sem & yea taken), 5 (approved status), 8 (approver name),
            9 (expiration date), 12 (comment)
        column 8 (approver name) is replaced with approver ID in approvers
        course ID from courses and major requirement ID from major_reqs
    majore_ws: worksheet
    major_id: index of worksheet in the workbook incremented by 1
    """
    evals = []
    for major_ws in transfer_wb:
        major_id = transfer_wb.sheetnames.index(major_ws.title) + 1
        for row in major_ws.iter_rows(
            min_row=2, max_row=major_ws.max_row, min_col=start, max_col=end,
            values_only=True):
            eval_data = eval_with_fk(
                major_id, row, schools, courses, approvers, major_reqs)
            evals.append(eval_data)
    #print(courses,'This is in evals')
    return evals


def get_data_by_major(
        transfer_wb, schools, courses, approvers, major_reqs, evals):
    """
    Computes schools list with unique school names from major_ws
    Computes courses list with unique data composed of
        3-element sublists that have:
            school ID, subject number, and course title
    Computes approvers list with unique approver names from major_ws
    Computes major-reqs list with unique data composed of
        2-element sublists that have major ID and description
    Computes transfer evalaution list with unique data composed of
        2-element sublists that have:
            course ID and major requirement ID
    major_ws: worksheet
    major_id: index of worksheet in the workbook incremented by 1
    schools,  courses, approvers, major_reqs, evals:
        Empty lists are passed into the call and extended with data from
        the worksheet.
    Returns:
        nothing. We use pass by reference mutable objects to update the
        parameters
    """
    school_col_idx = 1  # columh 'A'
    school_lst = get_unique_vals_from_col(transfer_wb, school_col_idx)
    schools.extend(school_lst)
    #print(schools)

    start_col = 1
    end_col = 3
    course_lst = get_course_by_major(transfer_wb, start_col, end_col, schools)
    #print(course_lst)
    for course in course_lst:
        if course not in courses:
            courses.append(course)
    #print(courses)

    approver_col_idx = 8  # column 'H'
    approver_lst = get_unique_vals_from_col(transfer_wb, approver_col_idx)
    approvers.extend(approver_lst)

    major_req_col_idx = 6  # column 'F'
    #print(major_id)
    major_descriptions = get_unique_reqs_vals_from_col(transfer_wb, major_req_col_idx)
    for description in major_descriptions:
        major_reqs.append(description)
    print(major_reqs,'major_reqs jutrfh')

    start_col = 1
    end_col = 12
    transfer_eval_lst = get_eval_by_major(
        transfer_wb, start_col, end_col, schools, courses, approvers,
        major_reqs)
    evals.extend(transfer_eval_lst)


def update(all_lst, one_lst):
    """
    Adds elements in one_lst to all_lst only if they are unique
    """
    for val in one_lst:
        if val not in all_lst:
            all_lst.append(val)
    return all_lst


def get_all_data(
        transfer_wb, all_majors, all_schools, all_courses, all_approvers,
        all_major_reqs, all_evals):
    """
    Builds all_majors list of major names
    Builds all_scholls list of school names
    Builds all_approvers list of names
    Builds all_major_reqs list of 2-element sublists composed of:
        major ID and description
    Builds all_courses list of 3-element sublists composed of:
        school ID, subject_num, and title
    Builds all_evals list of 7-element sublists composed of:
        course Id, majore requirement ID, semester/year taken, approved status,
        approver ID, expiration date, and comment
    transfer_wb: workbook
    all_majors, all_schools, all_approvers, all_major_reqs, all_courses,
    all_evals: lists, passed in as empty and built to have all the data for the
    database
    """
    # Iterate through all_major worksheets to get access to data on each
    # worksheet
    for idx, major_name in enumerate(all_majors):
        major_ws = transfer_wb[major_name]
        major_id = idx + 1
        schools = []
        approvers = []
        major_reqs = []
        courses = []
        evals = []
        # Passing empty lists and getting back lists with data for each entity
        get_data_by_major(
            major_ws, major_id, schools, courses, approvers, major_reqs, evals
        )
        # Use the worksheet lists to update build workbook lists
        update(all_schools, schools)
        update(all_approvers, approvers)
        update(all_major_reqs, major_reqs)
        update(all_courses, courses)
        update(all_evals, evals)


def test_get_all_data(transfer_wb):
    """
    Tests get all data from the workbook 'transfer_by_major.xlsx'
    """
    all_majors = transfer_wb.sheetnames
    all_schools = []
    all_major_reqs = []
    all_approvers = []
    all_courses = []
    all_evals = []
    get_all_data(
        transfer_wb,
        all_majors, all_schools, all_courses, all_approvers, all_major_reqs,
        all_evals
    )
    print('\nAll Majors\n==============')
    print(all_majors)

    # Arguments schools, major_reqs, approvers, and courses get updated after
    # the call
    print('\nFirst 10 Schools\n==========')
    print(all_schools[:10])
    print('\nFirst 10 Major Requirements\n===============')
    print(all_major_reqs[:50])
    print('\nAll Approvers\n==========')
    print(all_approvers)
    print('\nFirst 10 Courses\n==========')
    print(all_courses[:10])
    print('\nFirst 10 Transfer Evaluations\n=====================')
    print(all_evals[:10])
